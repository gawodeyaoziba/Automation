import threading
import time
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

# 文件路径模板
file_path_template = "测试报告_{thread_id}_{datetime_str}.xlsx"

def initialize_excel(file_path):
    """初始化 Excel 文件，创建标题行"""
    wb = Workbook()
    ws = wb.active
    ws.title = "测试报告"
    ws.append(["ID", "用例名称", "执行结果", "执行时间", "耗时时间"])
    wb.save(file_path)
    print(f"Excel 文件 '{file_path}' 初始化成功！")

def write_data_to_excel(file_path, id, name, report, times, timess):
    """将数据写入 Excel 文件"""
    if not os.path.exists(file_path):
        initialize_excel(file_path)

    wb = load_workbook(file_path)
    ws = wb.active

    row = [
        id,
        name,
        report,
        times.strftime('%Y-%m-%d %H:%M:%S'),
        timess
    ]
    ws.append(row)

    wb.save(file_path)
    print(f"数据写入 Excel 文件 '{file_path}' 成功！")

def thread_function(thread_id):
    """线程执行函数"""
    # 生成唯一的文件路径
    file_path = file_path_template.format(
        thread_id=thread_id,
        datetime_str=datetime.now().strftime('%Y%m%d_%H%M%S')
    )

    # 示例数据
    for i in range(3):
        id = i + 1
        name = f"用例{thread_id}_{id}"
        report = "成功"
        times = datetime.now()
        timess = "123.45 毫秒"

        write_data_to_excel(file_path, id, name, report, times, timess)
        time.sleep(5)  # 模拟处理时间

# 启动多个线程
threads = []
for i in range(3):  # 启动 3 个线程
    thread = threading.Thread(target=thread_function, args=(i,))
    threads.append(thread)
    thread.start()

# 等待所有线程完成
for thread in threads:
    thread.join()
