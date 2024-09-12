from filelock import FileLock, Timeout
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os
from Utils.ConfigOperate import ConfigOperate
report_file = ConfigOperate().ConfigContent(key="report", section="Path")

# 创建唯一的文件路径和锁文件路径


def initialize_excel(file_path, logger):
    """初始化 Excel 文件，创建标题行"""
    wb = Workbook()
    ws = wb.active
    ws.title = "测试报告"
    ws.append(["ID", "用例名称", "执行结果", "执行时间", "耗时时间"])
    wb.save(file_path)
    logger.info(f"Excel 文件 '{file_path}' 初始化成功！")


def Report(id, name, report, times, timess, logger, file_path):
    """将数据写入 Excel 文件"""
    if not os.path.exists(file_path):
        initialize_excel(file_path, logger)

    wb = load_workbook(file_path)
    ws = wb.active

    row = [
        id,
        name,
        report,
        times.strftime('%Y-%m-%d %H:%M:%S'),
        timess
    ]
    ws.append(row)

    wb.save(file_path)
    logger.info(f"数据写入 Excel 文件 '{file_path}' 成功！")


def write_result_to_excel(case_id, result, total_time_ms, file_path, logger):
    """
    将断言结果和耗时时间写入Excel文件
    :param case_id: 用例的ID (与Excel中行号对应)
    :param result: 断言结果（成功/失败）
    :param total_time_ms: 总耗时时间
    :param file_path: Excel文件路径
    :param logger: 日志记录器
    """
    try:
        workbook = load_workbook(file_path)
        sheet = workbook.active
        row = int(case_id) + 1

        result_to_record = "执行失败" if "断言失败" in result else "执行成功"
        sheet[f'H{row}'] = result_to_record
        sheet[f'I{row}'] = f'{total_time_ms:.2f} 毫秒'

        workbook.save(file_path)
        logger.info(f'用例 {case_id} 的执行结果和耗时时间已写入Excel文件')
    except Exception as e:
        logger.error(f'写入Excel文件失败: {e}')